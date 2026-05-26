"""Pyodide bridge module for Regression Analysis web app.

This module runs inside the Pyodide (WebAssembly) Python runtime.
It receives data and commands from JavaScript, and returns
JSON-serializable results.

Architecture:
    JS (HTML UI) -> pyodide.globals -> bridge functions -> JSON response

Packages required via pyodide.loadPackage:
    numpy, pandas, statsmodels, scipy, plotly, openpyxl
"""

from __future__ import annotations

import io
import json
import sys
import warnings
from typing import Any

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")


# ===========================================================================
# Utility: categorical detection
# ===========================================================================


def _detect_categorical_columns(df: pd.DataFrame) -> dict[str, str]:
    """Classify each column as 'numeric', 'categorical', or 'id'."""
    types: dict[str, str] = {}
    nrows = len(df)

    for col in df.columns:
        series = df[col]
        n_unique = series.nunique()

        # ID columns (unique values == nrows)
        col_lower = str(col).lower()
        id_patterns = ("id", "code", "num", "no.", "number", "序号", "编号", "代码")
        is_id_name = any(col_lower.startswith(p) or col_lower.endswith(p) for p in id_patterns)

        if is_id_name and n_unique == nrows and nrows > 0:
            types[str(col)] = "id"
            continue

        if pd.api.types.is_numeric_dtype(series):
            types[str(col)] = "numeric"
        elif pd.api.types.is_bool_dtype(series):
            types[str(col)] = "categorical"
        elif pd.api.types.is_object_dtype(series) or pd.api.types.is_string_dtype(series):
            # Try numeric conversion
            try:
                pd.to_numeric(series.dropna())
                types[str(col)] = "numeric"
            except (ValueError, TypeError):
                if n_unique / max(nrows, 1) > 0.9:
                    types[str(col)] = "id"
                else:
                    types[str(col)] = "categorical"
        else:
            types[str(col)] = "categorical"

    return types


# ===========================================================================
# 1. File parsing
# ===========================================================================


def parse_file(filename: str, content_b64: str) -> str:
    """Parse an uploaded CSV or Excel file and return column info + data.

    Args:
        filename: Original filename (used to detect format).
        content_b64: Base64-encoded file bytes.

    Returns:
        JSON string with keys:
            - success: bool
            - columns: list of {name, dtype, col_type, n_unique, n_missing, missing_rate}
            - data: list of lists (first row = headers)
            - n_rows, n_cols: int
            - error: str (if success=False)
    """
    import base64

    try:
        content_bytes = base64.b64decode(content_b64)
    except Exception as e:
        return json.dumps({"success": False, "error": f"Base64 decode failed: {e}"})

    name_lower = filename.lower()

    try:
        if (name_lower.endswith(".csv")
                or name_lower.endswith(".tsv")
                or name_lower.endswith(".txt")):
            # Detect encoding and separator
            if name_lower.endswith(".tsv"):
                sep = "\t"
            elif name_lower.endswith(".txt"):
                # Auto-detect separator for .txt: count tabs vs commas in first few lines
                sep = _detect_separator(content_bytes)
            else:
                sep = ","
            # Try UTF-8 first, then GBK
            for enc in ["utf-8", "gbk", "latin-1"]:
                try:
                    df = pd.read_csv(io.BytesIO(content_bytes), sep=sep, encoding=enc)
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            else:
                return json.dumps({"success": False, "error": "Cannot decode file encoding."})
        elif name_lower.endswith(".xls") and not name_lower.endswith(".xlsx"):
            # Old .xls format: try xlrd first, then fallback to openpyxl
            try:
                df = pd.read_excel(io.BytesIO(content_bytes), engine="xlrd")
            except (ImportError, Exception):
                try:
                    df = pd.read_excel(io.BytesIO(content_bytes), engine="openpyxl")
                except Exception as e2:
                    return json.dumps({"success": False, "error": f"Excel parse error: {e2}"})
        elif name_lower.endswith(".xlsx"):
            df = pd.read_excel(io.BytesIO(content_bytes), engine="openpyxl")
        else:
            return json.dumps({
                "success": False,
                "error": f"Unsupported format: {filename}. Use .csv, .tsv, .xls, or .xlsx.",
            })
    except Exception as e:
        return json.dumps({"success": False, "error": f"Parse error: {e}"})

    if df.empty:
        return json.dumps({"success": False, "error": "File is empty or has no data rows."})

    n_rows, n_cols = df.shape

    # Build column info
    col_types = _detect_categorical_columns(df)
    columns_info = []
    for col in df.columns:
        series = df[col]
        n_missing = int(series.isna().sum())
        columns_info.append({
            "name": str(col),
            "dtype": str(series.dtype),
            "col_type": col_types.get(str(col), "categorical"),
            "n_unique": int(series.nunique()),
            "n_missing": n_missing,
            "missing_rate": round(n_missing / max(n_rows, 1), 4),
        })

    # Data as list of lists (header row first)
    data_list = [list(df.columns)]
    for _, row in df.iterrows():
        data_list.append([_safe_value(v) for v in row])

    return json.dumps({
        "success": True,
        "columns": columns_info,
        "data": data_list,
        "n_rows": n_rows,
        "n_cols": n_cols,
    })


def _detect_separator(content_bytes: bytes) -> str:
    """Auto-detect CSV separator by counting tabs vs commas in first few lines."""
    sample = content_bytes[:4096].decode("utf-8", errors="ignore")
    lines = sample.splitlines()[:10]
    n_tabs = sum(line.count("\t") for line in lines if line.strip())
    n_commas = sum(line.count(",") for line in lines if line.strip())
    return "\t" if n_tabs > n_commas else ","


def _safe_value(v: Any) -> Any:
    """Convert numpy/pandas types to JSON-safe Python types."""
    if pd.isna(v):
        return None
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return float(v)
    if isinstance(v, (np.bool_,)):
        return bool(v)
    if isinstance(v, np.ndarray):
        return v.tolist()
    return v


# ===========================================================================
# 2. Regression engine
# ===========================================================================


def _validate_columns_metadata(columns_meta: Any, df: pd.DataFrame) -> bool:
    """Validate that columns metadata is a non-empty list with required fields.

    Checks that each entry has 'name' and 'col_type' fields, and that the
    names match actual DataFrame column names.

    Args:
        columns_meta: The columns metadata list (or None/missing).
        df: The DataFrame to validate against.

    Returns:
        True if metadata is valid and usable for dtype restoration.
    """
    if not isinstance(columns_meta, list) or len(columns_meta) == 0:
        return False
    df_cols = set(df.columns)
    for entry in columns_meta:
        if not isinstance(entry, dict):
            return False
        if "name" not in entry or "col_type" not in entry:
            return False
        if entry["name"] not in df_cols:
            return False
    return True


def _apply_data_filter(df: pd.DataFrame, filter_spec: dict) -> pd.DataFrame:
    """Apply row-subsetting filter to DataFrame.

    Args:
        df: DataFrame to filter.
        filter_spec: {col, type: 'numeric'|'categorical', min, max, values}

    Returns:
        Filtered DataFrame.
    """
    col = filter_spec.get("col")
    if not col or col not in df.columns:
        return df
    ftype = filter_spec.get("type", "")
    if ftype == "numeric":
        min_v = filter_spec.get("min")
        max_v = filter_spec.get("max")
        s = pd.to_numeric(df[col], errors="coerce")
        if min_v is not None:
            df = df[s >= min_v]
        if max_v is not None:
            df = df[s <= max_v]
    elif ftype == "categorical":
        values = filter_spec.get("values", [])
        if values:
            df = df[df[col].astype(str).isin(values)]
    return df


def _infer_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Convert columns that appear to be numeric from object/string dtype.

    For each column, attempts pd.to_numeric coercion. If all original
    non-null values successfully convert to numbers (no new NaN introduced),
    the column is replaced with the numeric version.

    Args:
        df: DataFrame potentially with object-dtype columns.

    Returns:
        The same DataFrame with numeric columns converted in-place.
    """
    for col in df.columns:
        col_series = df[col]
        nonnull_mask = col_series.notna()
        if nonnull_mask.sum() == 0:
            continue
        try:
            converted = pd.to_numeric(col_series, errors="coerce")
        except Exception:
            continue
        # Only convert if no original non-null values became NaN
        if (nonnull_mask & converted.isna()).sum() == 0:
            df[col] = converted
    return df


def run_regression(data_json: str, spec_json: str) -> str:
    """Run OLS or Logit regression.

    Args:
        data_json: JSON string with 'data' (list of lists) or 'columns'+'rows'.
        spec_json: JSON string with:
            - dep_var: str
            - indep_vars: list of str
            - has_intercept: bool (default true)
            - alpha: float (default 0.05)
            - cov_type: str (default 'nonrobust')
            - missing_strategy: str (default 'drop')

    Returns:
        JSON string with ModelResult, or error.
    """
    try:
        data_dict = json.loads(data_json)
        spec_dict = json.loads(spec_json)
    except json.JSONDecodeError as e:
        return json.dumps({"success": False, "error": f"JSON parse error: {e}"})

    # Reconstruct DataFrame
    try:
        if "data" in data_dict and isinstance(data_dict["data"], list):
            # Format: {data: [[header...], [row1...], ...]}
            rows = data_dict["data"]
            if len(rows) < 2:
                return json.dumps({"success": False, "error": "Data has no rows."})
            headers = rows[0]
            df = pd.DataFrame(rows[1:], columns=headers)
            # Convert numeric columns back from object dtype (JSON round-trip)
            if _validate_columns_metadata(data_dict.get("columns"), df):
                for col_info in data_dict["columns"]:
                    if (col_info.get("col_type") == "numeric"
                            and isinstance(col_info.get("name"), str)
                            and col_info["name"] in df.columns):
                        df[col_info["name"]] = pd.to_numeric(df[col_info["name"]], errors="coerce")
            else:
                print(
                    "[bridge] columns metadata missing or invalid, "
                    "using dtype inference fallback",
                    file=sys.stderr,
                )
                df = _infer_numeric_columns(df)
        elif "columns" in data_dict and "rows" in data_dict:
            df = pd.DataFrame(data_dict["rows"], columns=data_dict["columns"])
        else:
            return json.dumps({"success": False, "error": "Invalid data format."})
    except Exception as e:
        return json.dumps({"success": False, "error": f"DataFrame construction error: {e}"})

    dep_var = spec_dict.get("dep_var", "")
    indep_vars = spec_dict.get("indep_vars", [])
    has_intercept = spec_dict.get("has_intercept", True)
    alpha = spec_dict.get("alpha", 0.05)
    cov_type = spec_dict.get("cov_type", "nonrobust")
    missing_strategy = spec_dict.get("missing_strategy", "drop")
    model_type = spec_dict.get("model_type", "ols").lower()

    if not dep_var or not indep_vars:
        return json.dumps({"success": False, "error": "Must specify dep_var and indep_vars."})

    if dep_var not in df.columns:
        return json.dumps({"success": False, "error": f"dep_var '{dep_var}' not in data."})
    for v in indep_vars:
        if v not in df.columns:
            return json.dumps({"success": False, "error": f"Variable '{v}' not in data."})

    # --- Apply data filter (row subsetting) ---
    filter_spec = spec_dict.get("filter")
    if filter_spec:
        try:
            df = _apply_data_filter(df, filter_spec)
        except Exception as e:
            return json.dumps({"success": False, "error": f"Data filter error: {e}"})

    # Drop rows with missing values in relevant columns
    cols_used = [dep_var] + indep_vars
    df_clean = df[cols_used].copy()

    if missing_strategy == "drop":
        df_clean = df_clean.dropna()
    elif missing_strategy in ("mean", "median"):
        for col in indep_vars:
            if df_clean[col].isna().any():
                try:
                    fill_val = (
                        df_clean[col].mean()
                        if missing_strategy == "mean"
                        else df_clean[col].median()
                    )
                    df_clean[col] = df_clean[col].fillna(fill_val)
                except Exception:
                    # Non-numeric column: fall back to mode
                    mode_val = df_clean[col].mode()
                    if len(mode_val) > 0:
                        df_clean[col] = df_clean[col].fillna(mode_val[0])
                    else:
                        df_clean = df_clean.dropna(subset=[col])

            # Also handle dep_var missing values
            # For logit models, use mode fill for the DV to avoid corrupting binary values
            if df_clean[dep_var].isna().any():
                if model_type in ("logit", "probit"):
                    # Binary DV: fill with mode (most frequent class) instead of mean/median
                    mode_vals = df_clean[dep_var].mode()
                    if len(mode_vals) > 0:
                        df_clean[dep_var] = df_clean[dep_var].fillna(mode_vals[0])
                    else:
                        df_clean = df_clean.dropna(subset=[dep_var])
                else:
                    try:
                        fill_val = (
                            df_clean[dep_var].mean()
                            if missing_strategy == "mean"
                            else df_clean[dep_var].median()
                        )
                        df_clean[dep_var] = df_clean[dep_var].fillna(fill_val)
                    except Exception:
                        df_clean = df_clean.dropna(subset=[dep_var])

    if len(df_clean) < 2:
        return json.dumps({
            "success": False,
            "error": "Not enough valid observations after handling missing values.",
        })

    # --- Apply variable transformations ---
    transforms = spec_dict.get("transforms", {})
    var_name_map: dict[str, str] = {}  # original -> transformed column name
    if transforms:
        for var, ttype in transforms.items():
            if var not in df_clean.columns:
                return json.dumps({
                    "success": False,
                    "error": f"Transform variable '{var}' not in data.",
                })
            serie = pd.to_numeric(df_clean[var], errors="coerce")
            if ttype == "log":
                new_col = f"{var}_log"
                df_clean[new_col] = np.log(serie.clip(lower=1e-10))
            elif ttype == "standardize":
                new_col = f"{var}_z"
                sd = float(serie.std())
                if sd == 0:
                    df_clean[new_col] = 0.0
                else:
                    df_clean[new_col] = (serie - float(serie.mean())) / sd
            elif ttype == "center":
                new_col = f"{var}_c"
                df_clean[new_col] = serie - float(serie.mean())
            elif ttype == "square":
                new_col = f"{var}_sq"
                df_clean[new_col] = serie ** 2
            else:
                return json.dumps({
                    "success": False,
                    "error": f"Unsupported transform type: {ttype}",
                })
            var_name_map[var] = new_col
            # Replace var with transformed column in indep_vars list
            try:
                idx = indep_vars.index(var)
                indep_vars[idx] = new_col
            except ValueError:
                pass

    # --- Validate interaction terms ---
    interactions_list = spec_dict.get("interactions", [])
    if interactions_list:
        for pair in interactions_list:
            v1, v2 = pair[0], pair[1]
            # Resolve to actual column names (may have been transformed)
            actual_v1 = var_name_map.get(v1, v1)
            actual_v2 = var_name_map.get(v2, v2)
            if actual_v1 not in df_clean.columns:
                return json.dumps({
                    "success": False,
                    "error": (
                        f"Interaction variable '{v1}' "
                        "not in data after transforms."
                    ),
                })
            if actual_v2 not in df_clean.columns:
                return json.dumps({
                    "success": False,
                    "error": (
                        f"Interaction variable '{v2}' "
                        "not in data after transforms."
                    ),
                })

    # Build design matrix (interactions handled inside, matching patsy structure)
    try:
        X, y, coef_names, transform_map = _build_design_matrix(  # noqa: N806
            df_clean, dep_var, indep_vars, has_intercept,
            interactions=interactions_list,
        )
    except Exception as e:
        return json.dumps({"success": False, "error": f"Design matrix error: {e}"})

    # Fit model — dispatch to appropriate statsmodels class
    try:
        import statsmodels.api as sm

        if model_type == "logit":
            y_unique = np.unique(y)
            if len(y_unique) != 2:
                return json.dumps({
                    "success": False,
                    "error": (
                        f"Logit requires a binary dependent variable. "
                        f"Found {len(y_unique)} unique values "
                        f"in '{dep_var}': {list(y_unique)[:10]}."
                    ),
                })
            fitted = sm.Logit(y, X).fit(disp=False)

        elif model_type == "probit":
            y_unique = np.unique(y)
            if len(y_unique) != 2:
                return json.dumps({
                    "success": False,
                    "error": (
                        f"Probit requires a binary dependent variable. "
                        f"Found {len(y_unique)} unique values "
                        f"in '{dep_var}': {list(y_unique)[:10]}."
                    ),
                })
            fitted = sm.Probit(y, X).fit(disp=False)

        elif model_type in ("poisson", "negbin"):
            # Validate count data requirements
            if (y < 0).any():
                return json.dumps({
                    "success": False,
                    "error": f"{model_type.capitalize()} requires non-negative "
                             f"dependent variable. Found negative values in '{dep_var}'."
                })
            if not np.allclose(y, np.round(y), atol=1e-8):
                return json.dumps({
                    "success": False,
                    "error": f"{model_type.capitalize()} requires integer-valued "
                             f"dependent variable (count data)."
                })
            if model_type == "poisson":
                family = sm.families.Poisson()
            else:
                family = sm.families.NegativeBinomial()
            fitted = sm.GLM(y, X, family=family).fit()

        elif model_type == "mixedlm":
            # Need groups from spec
            group_col = spec_dict.get("group_var", "")
            if not group_col or group_col not in df.columns:
                return json.dumps({
                    "success": False,
                    "error": "MixedLM requires a valid 'group_var' in the spec."
                })
            # Align groups with rows that survived cleaning
            groups = df.loc[df_clean.index, group_col].values
            fitted = sm.MixedLM(y, X, groups=groups).fit(reml=True, disp=False)

        elif model_type == "panel":
            try:
                from linearmodels.panel import PanelOLS, RandomEffects
            except ImportError:
                return json.dumps({
                    "success": False,
                    "error": "Panel models require linearmodels package. "
                             "Not available in this Pyodide environment."
                })
            entity_col = spec_dict.get("entity_var", "")
            time_col = spec_dict.get("time_var", "")
            if not entity_col or entity_col not in df.columns:
                return json.dumps({
                    "success": False,
                    "error": "Panel model requires a valid 'entity_var' in the spec."
                })
            if not time_col or time_col not in df.columns:
                return json.dumps({
                    "success": False,
                    "error": "Panel model requires a valid 'time_var' in the spec."
                })
            # Build panel index
            valid_rows = df_clean.index
            entity_vals = df.loc[valid_rows, entity_col].values
            time_vals = df.loc[valid_rows, time_col].values
            panel_idx = pd.MultiIndex.from_arrays(
                [entity_vals, time_vals], names=[entity_col, time_col]
            )
            X_panel = pd.DataFrame(X, index=panel_idx)  # noqa: N806
            y_panel = pd.Series(y, index=panel_idx)
            panel_model_type = spec_dict.get("panel_model", "fixed")
            if panel_model_type == "random":
                model = RandomEffects(y_panel, X_panel)
            else:
                model = PanelOLS(y_panel, X_panel, entity_effects=True)
            cov_type_spec = cov_type if cov_type and cov_type != "nonrobust" else None
            if cov_type == "clustered" or cov_type_spec is None:
                fitted = model.fit(cov_type="clustered", cluster_entity=True)
            else:
                fitted = model.fit(cov_type=cov_type_spec)

        else:
            # OLS (default)
            if cov_type and cov_type != "nonrobust":
                fitted = sm.OLS(y, X).fit(cov_type=cov_type)
            else:
                fitted = sm.OLS(y, X).fit()

    except Exception as e:
        return json.dumps({"success": False, "error": f"Fit error: {e}"})

    # Extract results — dispatch to appropriate extractor
    if model_type == "logit":
        return _extract_logit_result(fitted, dep_var, indep_vars,
                                     coef_names, has_intercept, alpha,
                                     transform_map, df_clean)
    elif model_type == "probit":
        return _extract_probit_result(fitted, dep_var, indep_vars,
                                      coef_names, has_intercept, alpha,
                                      transform_map, df_clean)
    elif model_type in ("poisson", "negbin"):
        return _extract_count_result(fitted, dep_var, indep_vars,
                                     coef_names, has_intercept, alpha,
                                     transform_map, df_clean)
    elif model_type == "mixedlm":
        return _extract_mixedlm_result(fitted, dep_var, indep_vars,
                                       coef_names, has_intercept, alpha,
                                       transform_map, df_clean,
                                       spec_dict)
    elif model_type == "panel":
        return _extract_panel_result(fitted, dep_var, indep_vars,
                                     coef_names, has_intercept, alpha,
                                     transform_map, df_clean,
                                     spec_dict)
    else:
        return _extract_model_result(fitted, df_clean, dep_var, indep_vars,
                                     coef_names, has_intercept, alpha, cov_type,
                                     transform_map)


def _build_design_matrix(
    df: pd.DataFrame,
    dep_var: str,
    indep_vars: list[str],
    has_intercept: bool,
    interactions: list[tuple[str, str]] | None = None,
) -> tuple[np.ndarray, np.ndarray, list[str], dict[str, list[str]]]:
    """Build design matrix X and response vector y.

    Handles categorical variables by creating dummy variables via
    ``pd.get_dummies(drop_first=True)`` and optionally creates
    interaction columns for categorical-numeric and categorical-categorical
    pairs, matching the structure that patsy would produce.

    Args:
        df: Cleaned DataFrame with all needed columns.
        dep_var: Name of dependent variable column.
        indep_vars: List of independent variable column names.
        has_intercept: Whether to include an intercept column.
        interactions: Optional list of ``(var1, var2)`` interaction pairs.

    Returns:
        (X, y, coef_names, transform_map)
    """
    dep_series = df[dep_var]
    if not pd.api.types.is_numeric_dtype(dep_series):
        uniq = dep_series.dropna().unique()
        if len(uniq) == 2:
            mapping = {uniq[0]: 0, uniq[1]: 1}
            dep_series = dep_series.map(mapping)
        else:
            dep_series = pd.to_numeric(dep_series, errors="coerce")
    y = dep_series.astype(float).values

    interactions = interactions or []

    # ------------------------------------------------------------------
    # Step 1: Build main-effect columns for each independent variable
    # ------------------------------------------------------------------
    var_columns: dict[str, tuple[list[str], np.ndarray]] = {}

    for var in indep_vars:
        series = df[var]
        if pd.api.types.is_numeric_dtype(series):
            vals = series.astype(float).values.reshape(-1, 1)
            names = [var]
        else:
            # Categorical: one-hot encode, drop first to avoid dummy trap
            dummies = pd.get_dummies(series, prefix=var, drop_first=True, dtype=float)
            names = list(dummies.columns)
            vals = dummies.values
        var_columns[var] = (names, vals)

    # ------------------------------------------------------------------
    # Step 2: Build interaction columns from the main-effect dummies
    # ------------------------------------------------------------------
    interaction_columns: dict[tuple[str, str], tuple[list[str], np.ndarray]] = {}
    for v1, v2 in interactions:
        if v1 not in var_columns or v2 not in var_columns:
            raise ValueError(
                f"Interaction variable '{v1}' or '{v2}' not found in indep_vars."
            )
        names1, vals1 = var_columns[v1]
        names2, vals2 = var_columns[v2]

        int_names: list[str] = []
        int_vals_list: list[np.ndarray] = []
        for i, n1 in enumerate(names1):
            col1 = vals1[:, i]
            for j, n2 in enumerate(names2):
                col2 = vals2[:, j]
                int_name = f"{n1}:{n2}"
                int_val = col1 * col2
                int_names.append(int_name)
                int_vals_list.append(int_val)

        if int_vals_list:
            interaction_columns[(v1, v2)] = (
                int_names,
                np.column_stack(int_vals_list),
            )

    # ------------------------------------------------------------------
    # Step 3: Assemble the full design matrix
    # ------------------------------------------------------------------
    parts: list[np.ndarray] = []
    coef_names: list[str] = []
    transform_map: dict[str, list[str]] = {}

    for var in indep_vars:
        names, vals = var_columns[var]
        parts.append(vals)
        coef_names.extend(names)
        transform_map[var] = names

    for (v1, v2), (names, vals) in interaction_columns.items():
        parts.append(vals)
        coef_names.extend(names)
        int_key = f"{v1}:{v2}"
        transform_map[int_key] = names

    # Prepend intercept column (after main effects + interactions)
    if has_intercept:
        intercept_col = np.ones((len(y), 1))
        parts.insert(0, intercept_col)
        coef_names.insert(0, "Intercept")

    if not parts:
        raise ValueError("No predictor variables to build design matrix.")

    X = np.column_stack(parts)  # noqa: N806
    return X, y, coef_names, transform_map


def _build_variable_labels_for_web(
    coef_names: list[str],
    transform_map: dict[str, list[str]],
) -> dict[str, str]:
    """Build human-readable labels for coefficient names.

    For categorical dummy columns (e.g. ``education_B``) the label is
    ``education: B``.  Numeric columns and ``Intercept`` keep their
    original names.  Interaction terms (containing ``:``) are split into
    parts, each decoded individually, then joined with `` x ``.

    Args:
        coef_names: List of coefficient names from the design matrix.
        transform_map: Mapping from original variable name (or interaction
            key like ``"var1:var2"``) to the list of column names it
            produced in the design matrix.

    Returns:
        Dictionary mapping each coefficient name to its display label.
    """
    # Build reverse lookup: column_name -> original variable / interaction key
    col_to_var: dict[str, str] = {}
    for var_name, col_names in transform_map.items():
        for cname in col_names:
            col_to_var[cname] = var_name

    def _decode_single_part(part: str) -> str:
        """Decode one part of a coefficient name (main effect or interaction fragment)."""
        if part == "Intercept":
            return "Intercept"
        if part in col_to_var and col_to_var[part] != part:
            var_name = col_to_var[part]
            # The dummy column name is ``var_name + "_" + level``
            level = part[len(var_name) + 1:]  # +1 for the "_" separator
            return f"{var_name}: {level}"
        return part

    labels: dict[str, str] = {}
    for name in coef_names:
        if name == "Intercept":
            labels[name] = "Intercept"
        elif ":" in name:
            # Interaction term: split on ":", decode each fragment, rejoin
            fragments = name.split(":")
            decoded = [_decode_single_part(p) for p in fragments]
            labels[name] = " x ".join(decoded)
        elif name in col_to_var and col_to_var[name] != name:
            # Categorical dummy: extract level after the variable prefix
            var_name = col_to_var[name]
            level = name[len(var_name) + 1:]  # +1 for the "_" separator
            labels[name] = f"{var_name}: {level}"
        else:
            labels[name] = name

    return labels


def _extract_model_result(
    fitted,
    df: pd.DataFrame,
    dep_var: str,
    indep_vars: list[str],
    coef_names: list[str],
    has_intercept: bool,
    alpha: float,
    cov_type: str,
    transform_map: dict[str, list[str]],
) -> str:
    """Extract ModelResult from fitted statsmodels OLS and return JSON."""
    variable_labels = _build_variable_labels_for_web(coef_names, transform_map)

    params = np.asarray(fitted.params)
    bse = np.asarray(fitted.bse)
    tvalues = np.asarray(fitted.tvalues)
    pvalues = np.asarray(fitted.pvalues)
    conf_int = np.asarray(fitted.conf_int(alpha=alpha))

    coefficients = []
    for i, name in enumerate(coef_names):
        pv = float(pvalues[i])
        coefficients.append({
            "name": name,
            "coef": float(params[i]) if not np.isnan(params[i]) else None,
            "se": float(bse[i]) if not np.isnan(bse[i]) else None,
            "t_stat": float(tvalues[i]) if not np.isnan(tvalues[i]) else None,
            "pvalue": pv if not np.isnan(pv) else None,
            "ci_lower": float(conf_int[i, 0]) if not np.isnan(conf_int[i, 0]) else None,
            "ci_upper": float(conf_int[i, 1]) if not np.isnan(conf_int[i, 1]) else None,
            "significance": _significance_stars(pv),
        })

    n_obs = int(fitted.nobs)
    n_params = int(fitted.df_model) + (1 if has_intercept else 0)
    df_resid = int(fitted.df_resid)
    r_squared = float(fitted.rsquared) if hasattr(fitted, "rsquared") else None
    adj_r_squared = float(fitted.rsquared_adj) if hasattr(fitted, "rsquared_adj") else None

    f_stat = None
    if hasattr(fitted, "fvalue") and hasattr(fitted, "f_pvalue"):
        fv = float(fitted.fvalue)
        fp = float(fitted.f_pvalue)
        if not (np.isnan(fv) or np.isnan(fp)):
            f_stat = [fv, fp]

    log_likelihood = (
        float(fitted.llf)
        if hasattr(fitted, "llf") and fitted.llf is not None
        else None
    )
    aic = float(fitted.aic) if hasattr(fitted, "aic") else 0.0
    bic = float(fitted.bic) if hasattr(fitted, "bic") else 0.0
    ssr = float(fitted.ssr)
    rmse = float(np.sqrt(ssr / df_resid)) if df_resid > 0 else 0.0

    # Save residuals and fitted values for diagnostics
    residuals = fitted.resid.tolist() if hasattr(fitted, "resid") else []
    fitted_values = fitted.fittedvalues.tolist() if hasattr(fitted, "fittedvalues") else []

    # Model spec string
    preds_str = " + ".join(indep_vars)
    spec_str = f"{dep_var} ~ {preds_str}"
    if not has_intercept:
        spec_str += " - 1"
    if cov_type and cov_type != "nonrobust":
        spec_str += f" [SE: {cov_type}]"

    result = {
        "success": True,
        "model_type": "OLS",
        "coefficients": coefficients,
        "n_obs": n_obs,
        "n_params": n_params,
        "df_resid": df_resid,
        "r_squared": r_squared,
        "adj_r_squared": adj_r_squared,
        "f_statistic": f_stat,
        "log_likelihood": log_likelihood,
        "aic": aic,
        "bic": bic,
        "rmse": rmse,
        "dep_var": dep_var,
        "specification": spec_str,
        "method": "OLS",
        "se_type": cov_type if cov_type else "nonrobust",
        "residuals": residuals,
        "fitted_values": fitted_values,
        "indep_vars": indep_vars,
        "variable_labels": variable_labels,
    }

    return json.dumps(result)


def _extract_logit_result(
    fitted,
    dep_var: str,
    indep_vars: list[str],
    coef_names: list[str],
    has_intercept: bool,
    alpha: float,
    transform_map: dict[str, list[str]],
    df_clean: pd.DataFrame,
) -> str:
    """Extract logit regression results into JSON."""
    variable_labels = _build_variable_labels_for_web(coef_names, transform_map)

    params = np.asarray(fitted.params)
    bse = np.asarray(fitted.bse)
    zvalues = np.asarray(fitted.tvalues)  # statsmodels stores z as tvalues for Logit
    pvalues = np.asarray(fitted.pvalues)
    conf_int = np.asarray(fitted.conf_int(alpha=alpha))

    coefficients = []
    for i, name in enumerate(coef_names):
        pv = float(pvalues[i])
        coef_val = float(params[i]) if not np.isnan(params[i]) else None
        coefficients.append({
            "name": name,
            "coef": coef_val,
            "se": float(bse[i]) if not np.isnan(bse[i]) else None,
            "z_stat": float(zvalues[i]) if not np.isnan(zvalues[i]) else None,
            "pvalue": pv if not np.isnan(pv) else None,
            "ci_lower": float(conf_int[i, 0]) if not np.isnan(conf_int[i, 0]) else None,
            "ci_upper": float(conf_int[i, 1]) if not np.isnan(conf_int[i, 1]) else None,
            "odds_ratio": float(np.exp(coef_val)) if coef_val is not None else None,
            "or_ci_lower": float(np.exp(conf_int[i, 0])) if not np.isnan(conf_int[i, 0]) else None,
            "or_ci_upper": float(np.exp(conf_int[i, 1])) if not np.isnan(conf_int[i, 1]) else None,
            "significance": _significance_stars(pv),
        })

    n_obs = int(fitted.nobs)
    n_params = int(fitted.df_model) + (1 if has_intercept else 0)
    df_resid = int(fitted.df_resid)

    # McFadden's pseudo R-squared
    ll_model = float(fitted.llf) if hasattr(fitted, "llf") and fitted.llf is not None else 0.0
    ll_null = float(fitted.llnull) if hasattr(fitted, "llnull") else 0.0
    pseudo_r_squared = float(1.0 - ll_model / ll_null) if ll_null != 0 else None

    # Likelihood ratio test
    llr = float(fitted.llr) if hasattr(fitted, "llr") else None
    llr_pvalue = float(fitted.llr_pvalue) if hasattr(fitted, "llr_pvalue") else None

    log_likelihood = ll_model
    aic = float(fitted.aic) if hasattr(fitted, "aic") else 0.0
    bic = float(fitted.bic) if hasattr(fitted, "bic") else 0.0

    # Save residuals and fitted values for diagnostics
    residuals = fitted.resid_dev.tolist() if hasattr(fitted, "resid_dev") else []
    fitted_values = fitted.fittedvalues.tolist() if hasattr(fitted, "fittedvalues") else []

    # Model spec string
    preds_str = " + ".join(indep_vars)
    spec_str = f"{dep_var} ~ {preds_str}"
    if not has_intercept:
        spec_str += " - 1"

    # For logit: store design matrix data for ROC computation
    # We'll store predicted probabilities separately for efficient ROC generation
    y_actual = fitted.model.endog.tolist() if hasattr(fitted, "model") else []

    result = {
        "success": True,
        "model_type": "logit",
        "coefficients": coefficients,
        "n_obs": n_obs,
        "n_params": n_params,
        "df_resid": df_resid,
        "r_squared": None,
        "adj_r_squared": None,
        "pseudo_r_squared": pseudo_r_squared,
        "llr": llr,
        "llr_pvalue": llr_pvalue,
        "log_likelihood": log_likelihood,
        "aic": aic,
        "bic": bic,
        "rmse": None,
        "dep_var": dep_var,
        "specification": spec_str,
        "method": "Logit",
        "se_type": "MLE",
        "residuals": residuals,
        "fitted_values": fitted_values,
        "y_actual": y_actual,
        "indep_vars": indep_vars,
        "variable_labels": variable_labels,
    }

    return json.dumps(result)


def _extract_probit_result(
    fitted,
    dep_var: str,
    indep_vars: list[str],
    coef_names: list[str],
    has_intercept: bool,
    alpha: float,
    transform_map: dict[str, list[str]],
    df_clean: pd.DataFrame,
) -> str:
    """Extract probit regression results into JSON (no odds ratios)."""
    variable_labels = _build_variable_labels_for_web(coef_names, transform_map)

    params = np.asarray(fitted.params)
    bse = np.asarray(fitted.bse)
    zvalues = np.asarray(fitted.tvalues)
    pvalues = np.asarray(fitted.pvalues)
    conf_int = np.asarray(fitted.conf_int(alpha=alpha))

    coefficients = []
    for i, name in enumerate(coef_names):
        pv = float(pvalues[i])
        coefficients.append({
            "name": name,
            "coef": float(params[i]) if not np.isnan(params[i]) else None,
            "se": float(bse[i]) if not np.isnan(bse[i]) else None,
            "z_stat": float(zvalues[i]) if not np.isnan(zvalues[i]) else None,
            "pvalue": pv if not np.isnan(pv) else None,
            "ci_lower": float(conf_int[i, 0]) if not np.isnan(conf_int[i, 0]) else None,
            "ci_upper": float(conf_int[i, 1]) if not np.isnan(conf_int[i, 1]) else None,
            "significance": _significance_stars(pv),
        })

    n_obs = int(fitted.nobs)
    n_params = int(fitted.df_model) + (1 if has_intercept else 0)
    df_resid = int(fitted.df_resid)

    ll_model = float(fitted.llf) if hasattr(fitted, "llf") and fitted.llf is not None else 0.0
    ll_null = float(fitted.llnull) if hasattr(fitted, "llnull") else 0.0
    pseudo_r_squared = float(1.0 - ll_model / ll_null) if ll_null != 0 else None

    llr = float(fitted.llr) if hasattr(fitted, "llr") else None
    llr_pvalue = float(fitted.llr_pvalue) if hasattr(fitted, "llr_pvalue") else None

    log_likelihood = ll_model
    aic = float(fitted.aic) if hasattr(fitted, "aic") else 0.0
    bic = float(fitted.bic) if hasattr(fitted, "bic") else 0.0

    residuals = fitted.resid_dev.tolist() if hasattr(fitted, "resid_dev") else []
    fitted_values = fitted.fittedvalues.tolist() if hasattr(fitted, "fittedvalues") else []
    y_actual = fitted.model.endog.tolist() if hasattr(fitted, "model") else []

    preds_str = " + ".join(indep_vars)
    spec_str = f"{dep_var} ~ {preds_str}"
    if not has_intercept:
        spec_str += " - 1"

    result = {
        "success": True,
        "model_type": "probit",
        "coefficients": coefficients,
        "n_obs": n_obs,
        "n_params": n_params,
        "df_resid": df_resid,
        "r_squared": None,
        "adj_r_squared": None,
        "pseudo_r_squared": pseudo_r_squared,
        "llr": llr,
        "llr_pvalue": llr_pvalue,
        "log_likelihood": log_likelihood,
        "aic": aic,
        "bic": bic,
        "rmse": None,
        "dep_var": dep_var,
        "specification": spec_str,
        "method": "Probit",
        "se_type": "MLE",
        "residuals": residuals,
        "fitted_values": fitted_values,
        "y_actual": y_actual,
        "indep_vars": indep_vars,
        "variable_labels": variable_labels,
    }

    return json.dumps(result)


def _extract_count_result(
    fitted,
    dep_var: str,
    indep_vars: list[str],
    coef_names: list[str],
    has_intercept: bool,
    alpha: float,
    transform_map: dict[str, list[str]],
    df_clean: pd.DataFrame,
) -> str:
    """Extract Poisson/NegBin regression results into JSON (with IRR)."""
    variable_labels = _build_variable_labels_for_web(coef_names, transform_map)

    # Detect model subtype from GLM family
    family_name = getattr(fitted.family, "family", "")
    if "poisson" in family_name.lower():
        model_subtype = "poisson"
        method = "Poisson"
    else:
        model_subtype = "negbin"
        method = "NegativeBinomial"

    params = np.asarray(fitted.params)
    bse = np.asarray(fitted.bse)
    zvalues = np.asarray(fitted.tvalues)
    pvalues = np.asarray(fitted.pvalues)
    conf_int = np.asarray(fitted.conf_int(alpha=alpha))

    coefficients = []
    for i, name in enumerate(coef_names):
        pv = float(pvalues[i])
        coef_val = float(params[i]) if not np.isnan(params[i]) else None
        coefficients.append({
            "name": name,
            "coef": coef_val,
            "se": float(bse[i]) if not np.isnan(bse[i]) else None,
            "z_stat": float(zvalues[i]) if not np.isnan(zvalues[i]) else None,
            "pvalue": pv if not np.isnan(pv) else None,
            "ci_lower": float(conf_int[i, 0]) if not np.isnan(conf_int[i, 0]) else None,
            "ci_upper": float(conf_int[i, 1]) if not np.isnan(conf_int[i, 1]) else None,
            "irr": float(np.exp(coef_val)) if coef_val is not None else None,
            "significance": _significance_stars(pv),
        })

    n_obs = int(fitted.nobs)
    n_params = int(fitted.df_model) + (1 if has_intercept else 0)
    df_resid = int(fitted.df_resid)

    ll_model = float(fitted.llf) if hasattr(fitted, "llf") and fitted.llf is not None else 0.0
    try:
        ll_null = float(fitted.llnull)
    except (AttributeError, Exception):
        ll_null = 0.0
    pseudo_r_squared = float(1.0 - ll_model / ll_null) if ll_null != 0 else None

    # LLR from deviance
    llr = None
    llr_pvalue = None
    try:
        deviance = float(fitted.deviance)
        null_deviance = (
            float(fitted.null_deviance)
            if hasattr(fitted, "null_deviance")
            else deviance
        )
        if null_deviance > deviance:
            llr = float(null_deviance - deviance)
            df_llr = int(fitted.df_model)
            if llr > 0 and df_llr > 0:
                from scipy import stats as scipy_stats
                llr_pvalue = float(1.0 - scipy_stats.chi2.cdf(llr, df_llr))
    except Exception:
        pass

    log_likelihood = ll_model
    aic = float(fitted.aic) if hasattr(fitted, "aic") else 0.0
    # BIC: prefer llf-based
    bic = 0.0
    if hasattr(fitted, "bic_llf"):
        bic = float(fitted.bic_llf)
    elif hasattr(fitted, "bic"):
        bic = float(fitted.bic)

    # Dispersion for NegBin
    dispersion = None
    if model_subtype == "negbin":
        try:
            dispersion = float(fitted.scale)
        except Exception:
            pass

    residuals = fitted.resid_response.tolist() if hasattr(fitted, "resid_response") else []
    fitted_values = fitted.fittedvalues.tolist() if hasattr(fitted, "fittedvalues") else []

    preds_str = " + ".join(indep_vars)
    spec_str = f"{dep_var} ~ {preds_str}"
    if not has_intercept:
        spec_str += " - 1"

    result = {
        "success": True,
        "model_type": model_subtype,
        "coefficients": coefficients,
        "n_obs": n_obs,
        "n_params": n_params,
        "df_resid": df_resid,
        "r_squared": None,
        "adj_r_squared": None,
        "pseudo_r_squared": pseudo_r_squared,
        "llr": llr,
        "llr_pvalue": llr_pvalue,
        "log_likelihood": log_likelihood,
        "aic": aic,
        "bic": bic,
        "rmse": None,
        "dep_var": dep_var,
        "specification": spec_str,
        "method": method,
        "se_type": "MLE",
        "dispersion": dispersion,
        "residuals": residuals,
        "fitted_values": fitted_values,
        "indep_vars": indep_vars,
        "variable_labels": variable_labels,
    }

    return json.dumps(result)


def _extract_mixedlm_result(
    fitted,
    dep_var: str,
    indep_vars: list[str],
    coef_names: list[str],
    has_intercept: bool,
    alpha: float,
    transform_map: dict[str, list[str]],
    df_clean: pd.DataFrame,
    spec_dict: dict,
) -> str:
    """Extract MixedLM regression results into JSON."""
    variable_labels = _build_variable_labels_for_web(coef_names, transform_map)

    fe_names = fitted.fe_params.index
    params = fitted.fe_params
    bse = fitted.bse_fe
    tvalues = fitted.tvalues.loc[fe_names] if hasattr(fitted.tvalues, "loc") else fitted.tvalues
    pvalues = fitted.pvalues.loc[fe_names] if hasattr(fitted.pvalues, "loc") else fitted.pvalues
    conf_int = fitted.conf_int(alpha=alpha)

    coefficients = []
    for name in fe_names:
        pv = float(pvalues[name])
        conf_row = conf_int.loc[name] if hasattr(conf_int, "loc") else conf_int
        ci_low_val = float(conf_row[0])
        ci_high_val = float(conf_row[1])
        coefficients.append({
            "name": str(name),
            "coef": float(params[name]),
            "se": float(bse[name]),
            "t_stat": float(tvalues[name]),
            "pvalue": pv if not np.isnan(pv) else None,
            "ci_lower": ci_low_val if not np.isnan(ci_low_val) else None,
            "ci_upper": ci_high_val if not np.isnan(ci_high_val) else None,
            "significance": _significance_stars(pv),
        })

    n_obs = int(fitted.nobs)
    n_params = fitted.k_fe
    df_resid = int(fitted.df_resid)

    # R-squared from residuals
    y_endog = fitted.model.endog
    ss_resid = float(np.sum(fitted.resid ** 2))
    ss_total = float(np.sum((y_endog - y_endog.mean()) ** 2))
    r_squared = 1.0 - ss_resid / ss_total if ss_total > 0 else None
    adj_r_squared = (
        (1.0 - (1.0 - r_squared) * (n_obs - 1) / df_resid)
        if r_squared is not None and df_resid > 0
        else None
    )

    log_likelihood = (
        float(fitted.llf)
        if fitted.llf is not None and not np.isnan(fitted.llf)
        else None
    )

    aic = 0.0
    bic = 0.0
    if hasattr(fitted, "aic") and not np.isnan(fitted.aic):
        aic = float(fitted.aic)
    if hasattr(fitted, "bic") and not np.isnan(fitted.bic):
        bic = float(fitted.bic)

    rmse = float(np.sqrt(ss_resid / df_resid)) if df_resid > 0 else None

    # Random effects variance components
    re_var = {}
    if fitted.cov_re is not None and fitted.cov_re.size > 0:
        for i, name in enumerate(fitted.cov_re.index):
            re_var[str(name)] = float(fitted.cov_re.iloc[i, i])

    residuals = fitted.resid.tolist() if hasattr(fitted, "resid") else []
    fitted_values = fitted.fittedvalues.tolist() if hasattr(fitted, "fittedvalues") else []

    group_col = spec_dict.get("group_var", "unknown")
    group_count = 0
    if hasattr(fitted, "random_effects"):
        group_count = len(fitted.random_effects)

    preds_str = " + ".join(indep_vars)
    spec_str = f"{dep_var} ~ {preds_str}"
    if not has_intercept:
        spec_str += " - 1"
    spec_str += f"  [groups: {group_col} ({group_count})]"

    result = {
        "success": True,
        "model_type": "mixedlm",
        "coefficients": coefficients,
        "n_obs": n_obs,
        "n_params": n_params,
        "df_resid": df_resid,
        "r_squared": r_squared,
        "adj_r_squared": adj_r_squared,
        "pseudo_r_squared": None,
        "llr": None,
        "llr_pvalue": None,
        "log_likelihood": log_likelihood,
        "aic": aic,
        "bic": bic,
        "rmse": rmse,
        "dep_var": dep_var,
        "specification": spec_str,
        "method": "MixedLM (REML)",
        "se_type": "MixedLM",
        "group_var": group_col,
        "group_count": group_count,
        "re_var": re_var,
        "residuals": residuals,
        "fitted_values": fitted_values,
        "indep_vars": indep_vars,
        "variable_labels": variable_labels,
    }

    return json.dumps(result)


def _extract_panel_result(
    fitted,
    dep_var: str,
    indep_vars: list[str],
    coef_names: list[str],
    has_intercept: bool,
    alpha: float,
    transform_map: dict[str, list[str]],
    df_clean: pd.DataFrame,
    spec_dict: dict,
) -> str:
    """Extract panel regression results into JSON."""
    variable_labels = _build_variable_labels_for_web(coef_names, transform_map)

    is_fe = hasattr(fitted, "included_effects")
    panel_method = "Panel FE" if is_fe else "Panel RE"

    params = fitted.params
    std_errors = fitted.std_errors
    t_stat_vals = fitted.tstats
    p_vals = fitted.pvalues

    # CI
    try:
        conf_int = fitted.conf_int(alpha=alpha)
    except TypeError:
        conf_int = fitted.conf_int()
    ci_lower_col = "lower"
    ci_upper_col = "upper"

    coefficients = []
    for var_name in params.index:
        pv = float(p_vals[var_name])
        coefficients.append({
            "name": str(var_name),
            "coef": float(params[var_name]),
            "se": float(std_errors[var_name]),
            "t_stat": float(t_stat_vals[var_name]),
            "pvalue": pv if not np.isnan(pv) else None,
            "ci_lower": float(conf_int.loc[var_name, ci_lower_col]),
            "ci_upper": float(conf_int.loc[var_name, ci_upper_col]),
            "significance": _significance_stars(pv),
        })

    n_obs = int(fitted.nobs)
    n_params = int(fitted.df_model)
    df_resid = int(fitted.df_resid)

    # R-squared variants
    within_r2 = float(fitted.rsquared_within) if hasattr(fitted, "rsquared_within") else None
    between_r2 = float(fitted.rsquared_between) if hasattr(fitted, "rsquared_between") else None
    overall_r2 = float(fitted.rsquared_overall) if hasattr(fitted, "rsquared_overall") else None
    r_squared = within_r2 if within_r2 is not None else overall_r2

    log_likelihood = None
    try:
        if hasattr(fitted, "loglik") and fitted.loglik is not None:
            log_likelihood = float(fitted.loglik)
    except Exception:
        pass

    aic = float(fitted.aic) if hasattr(fitted, "aic") else 0.0
    bic = float(fitted.bic) if hasattr(fitted, "bic") else 0.0

    rmse = None
    if hasattr(fitted, "resid_ss") and df_resid > 0:
        rmse = float(np.sqrt(fitted.resid_ss / df_resid))

    # Entity / time counts
    n_entities = 0
    n_periods = 0
    try:
        n_entities = int(float(fitted.entity_info["total"]))
    except Exception:
        pass
    try:
        n_periods = int(float(fitted.time_info["total"]))
    except Exception:
        pass

    residuals = []
    fitted_values = []
    try:
        residuals = fitted.resids.values.flatten().tolist() if hasattr(fitted, "resids") else []
    except Exception:
        pass
    try:
        fitted_values = (
            fitted.fitted_values.values.flatten().tolist()
            if hasattr(fitted, "fitted_values")
            else []
        )
    except Exception:
        pass

    preds_str = " + ".join(indep_vars)
    spec_str = f"{dep_var} ~ {preds_str}"
    if not has_intercept:
        spec_str += " - 1"
    spec_str += f"  [{panel_method}]"

    result = {
        "success": True,
        "model_type": "panel",
        "coefficients": coefficients,
        "n_obs": n_obs,
        "n_params": n_params,
        "df_resid": df_resid,
        "r_squared": r_squared,
        "adj_r_squared": None,
        "pseudo_r_squared": None,
        "llr": None,
        "llr_pvalue": None,
        "log_likelihood": log_likelihood,
        "aic": aic,
        "bic": bic,
        "rmse": rmse,
        "dep_var": dep_var,
        "specification": spec_str,
        "method": panel_method,
        "se_type": "clustered",
        "within_r_squared": within_r2,
        "between_r_squared": between_r2,
        "overall_r_squared": overall_r2,
        "entity_count": n_entities,
        "time_count": n_periods,
        "panel_type": panel_method,
        "residuals": residuals,
        "fitted_values": fitted_values,
        "indep_vars": indep_vars,
        "variable_labels": variable_labels,
    }

    return json.dumps(result)


def _significance_stars(pvalue: float) -> str:
    if pvalue < 0.01:
        return "***"
    if pvalue < 0.05:
        return "**"
    if pvalue < 0.1:
        return "*"
    return ""


# ===========================================================================
# 3. Diagnostics
# ===========================================================================


def compute_diagnostics(data_json: str, result_json: str) -> str:
    """Compute diagnostic statistics from model results.

    Args:
        data_json: JSON data (same format as run_regression).
        result_json: JSON result from run_regression.

    Returns:
        JSON with VIF, residual tests, ANOVA table.
    """
    try:
        result = json.loads(result_json)
    except json.JSONDecodeError:
        return json.dumps({"success": False, "error": "Invalid result JSON."})

    residuals = np.array(result.get("residuals", []))
    n_obs = result.get("n_obs", 0)
    n_params = result.get("n_params", 0)
    r_squared = result.get("r_squared")
    rmse = result.get("rmse", 0.0)

    # --- VIF ---
    vif_df = _compute_vif(data_json, result)

    # --- Residual tests ---
    residual_diag = _compute_residual_tests(residuals)

    # --- ANOVA ---
    anova = _compute_anova(rmse, r_squared, n_obs, n_params, result)

    return json.dumps({
        "success": True,
        "vif": vif_df,
        "residual_tests": residual_diag,
        "anova": anova,
    })


def _compute_vif(data_json: str, result: dict) -> list[dict] | None:
    """Compute VIF for predictor variables."""
    try:
        from statsmodels.stats.outliers_influence import variance_inflation_factor
        from statsmodels.tools.tools import add_constant
    except ImportError:
        return None

    try:
        data_dict = json.loads(data_json) if isinstance(data_json, str) else data_json
        if "data" in data_dict:
            rows = data_dict["data"]
            if len(rows) < 2:
                return None
            df = pd.DataFrame(rows[1:], columns=rows[0])
            # Convert numeric columns back from object dtype (JSON round-trip)
            if _validate_columns_metadata(data_dict.get("columns"), df):
                for col_info in data_dict["columns"]:
                    if (col_info.get("col_type") == "numeric"
                            and isinstance(col_info.get("name"), str)
                            and col_info["name"] in df.columns):
                        df[col_info["name"]] = pd.to_numeric(df[col_info["name"]], errors="coerce")
            else:
                print(
                    "[bridge] columns metadata missing or invalid, "
                    "using dtype inference fallback",
                    file=sys.stderr,
                )
                df = _infer_numeric_columns(df)
        else:
            return None
    except Exception:
        return None

    indep_vars = result.get("indep_vars", [])
    numeric_vars = []
    for v in indep_vars:
        if v in df.columns and pd.api.types.is_numeric_dtype(df[v]):
            numeric_vars.append(v)

    if len(numeric_vars) < 2:
        return None

    try:
        X = df[numeric_vars].dropna().astype(float)  # noqa: N806
        X_c = add_constant(X)  # noqa: N806

        vif_rows = []
        for i in range(X_c.shape[1]):
            col_name = str(X_c.columns[i])
            vif_val = float(variance_inflation_factor(X_c.values, i))
            if np.isnan(vif_val) or np.isinf(vif_val):
                continue
            diagnosis = "High" if vif_val > 10 else "Moderate" if vif_val > 5 else "Low"
            vif_rows.append({
                "variable": col_name,
                "vif": round(vif_val, 4),
                "diagnosis": diagnosis,
            })
        return vif_rows
    except Exception:
        return None


def _compute_residual_tests(residuals: np.ndarray) -> dict:
    """Run Shapiro-Wilk and Durbin-Watson tests."""
    result = {}

    # Shapiro-Wilk
    if len(residuals) >= 3:
        try:
            from scipy import stats
            shapiro_stat, shapiro_p = stats.shapiro(residuals)
            result["shapiro_stat"] = round(float(shapiro_stat), 6)
            result["shapiro_pvalue"] = float(shapiro_p)
            result["shapiro_normal"] = "Yes" if shapiro_p > 0.05 else "No"
        except Exception:
            result["shapiro_normal"] = "Error"
    else:
        result["shapiro_normal"] = "Insufficient data"

    # Durbin-Watson
    if len(residuals) >= 2:
        try:
            diff_sum = np.sum(np.diff(residuals) ** 2)
            total_sum = np.sum(residuals ** 2)
            if total_sum > 0:
                dw = float(diff_sum / total_sum)
                result["dw_stat"] = round(dw, 4)
                if dw < 1.0:
                    result["dw_autocorrelation"] = "Positive (strong)"
                elif dw > 3.0:
                    result["dw_autocorrelation"] = "Negative (strong)"
                elif dw < 1.5:
                    result["dw_autocorrelation"] = "Positive (mild)"
                elif dw > 2.5:
                    result["dw_autocorrelation"] = "Negative (mild)"
                else:
                    result["dw_autocorrelation"] = "None"
            else:
                result["dw_autocorrelation"] = "N/A (zero variance)"
        except Exception:
            result["dw_autocorrelation"] = "Error"
    else:
        result["dw_autocorrelation"] = "Insufficient data"

    return result


def _compute_anova(rmse: float, r_squared: float | None,
                   n_obs: int, n_params: int, result: dict) -> dict:
    """Compute ANOVA table from model results."""
    df_resid = result.get("df_resid", max(n_obs - n_params, 1))
    ss_resid = rmse ** 2 * df_resid

    df_explained = n_params - 1 if n_params > 1 else 0
    df_total = n_obs - 1

    if r_squared is not None and r_squared < 1.0 and r_squared >= 0:
        ss_total = ss_resid / (1.0 - r_squared)
    elif r_squared is not None and r_squared >= 1.0:
        ss_total = ss_resid
    else:
        ss_total = ss_resid

    ss_explained = ss_total - ss_resid

    ms_explained = ss_explained / df_explained if df_explained > 0 else float("nan")
    ms_resid = ss_resid / df_resid if df_resid > 0 else float("nan")

    f_val = float("nan")
    f_p = float("nan")
    if ms_resid > 0 and ms_explained > 0:
        f_val = ms_explained / ms_resid
        try:
            from scipy import stats
            f_p = float(stats.f.sf(f_val, df_explained, df_resid))
        except Exception:
            pass

    f_stat = result.get("f_statistic")
    if f_stat and len(f_stat) == 2:
        f_val = f_stat[0]
        f_p = f_stat[1]

    return {
        "explained": {
            "source": "Regression",
            "SS": round(ss_explained, 6),
            "df": df_explained,
            "MS": round(ms_explained, 6) if not np.isnan(ms_explained) else None,
            "F": round(f_val, 6) if not np.isnan(f_val) else None,
            "p_value": round(f_p, 6) if not np.isnan(f_p) else None,
        },
        "residual": {
            "source": "Residual",
            "SS": round(ss_resid, 6),
            "df": df_resid,
            "MS": round(ms_resid, 6) if not np.isnan(ms_resid) else None,
        },
        "total": {
            "source": "Total",
            "SS": round(ss_total, 6),
            "df": df_total,
        },
    }


# ===========================================================================
# 4. Charts (plotly JSON)
# ===========================================================================


def generate_diagnostic_charts(result_json: str) -> str:
    """Generate diagnostic chart data from model results.

    Returns plotly-compatible JSON for:
        - residual_fitted
        - qq_plot
        - scale_location
        - cooks_distance
    """
    try:
        result = json.loads(result_json)
    except json.JSONDecodeError:
        return json.dumps({"success": False, "error": "Invalid JSON."})

    residuals = np.array(result.get("residuals", []))
    fitted_values = np.array(result.get("fitted_values", []))
    n_params = result.get("n_params", 0)

    charts = {}

    if len(residuals) >= 2 and len(fitted_values) >= 2:
        charts["residual_fitted"] = _make_residual_fitted_chart(residuals, fitted_values)
    else:
        charts["residual_fitted"] = None

    if len(residuals) >= 4:
        charts["qq"] = _make_qq_chart(residuals)
    else:
        charts["qq"] = None

    if len(residuals) >= 5:
        charts["scale_location"] = _make_scale_location_chart(residuals, fitted_values)
    else:
        charts["scale_location"] = None

    model_type = result.get("model_type", "")
    is_mle = model_type in ("logit", "probit", "poisson", "negbin")
    if len(residuals) >= 3 and not is_mle:
        charts["cooks_distance"] = _make_cooks_chart(residuals, fitted_values, n_params)
    elif is_mle:
        # Cook's distance formula (OLS-based) is not applicable to MLE model deviance residuals
        charts["cooks_distance"] = _make_unavailable_chart(
            "Cook's Distance",
            "Cook's distance is not applicable to MLE models. "
            "Consider using Pregibon's delta-beta influence statistic instead."
        )
    else:
        charts["cooks_distance"] = None

    return json.dumps({"success": True, "charts": charts})


def generate_coefficient_chart(result_json: str) -> str:
    """Generate a coefficient dot-whisker chart as plotly JSON."""
    try:
        result = json.loads(result_json)
    except json.JSONDecodeError:
        return json.dumps({"success": False, "error": "Invalid JSON."})

    coefficients = result.get("coefficients", [])
    if not coefficients:
        return json.dumps({"success": False, "error": "No coefficients."})

    # Separate intercept and sort others by absolute value
    intercept = None
    others = []
    for c in coefficients:
        if c["name"] == "Intercept":
            intercept = c
        else:
            others.append(c)
    others.sort(key=lambda x: abs(x.get("coef", 0)), reverse=True)

    sorted_coefs = []
    if intercept:
        sorted_coefs.append(intercept)
    sorted_coefs.extend(others)

    n = len(sorted_coefs)
    names = [c["name"] for c in sorted_coefs]
    estimates = [c["coef"] for c in sorted_coefs]
    ci_lows = [c["ci_lower"] for c in sorted_coefs]
    ci_highs = [c["ci_upper"] for c in sorted_coefs]

    # Build plotly figure spec
    traces = []
    for i in range(n):
        traces.append({
            "type": "scatter",
            "x": [ci_lows[i], ci_highs[i]],
            "y": [n - 1 - i, n - 1 - i],
            "mode": "lines",
            "line": {"color": "#1f77b4", "width": 2},
            "showlegend": False,
            "hoverinfo": "none",
        })

    traces.append({
        "type": "scatter",
        "x": estimates,
        "y": list(range(n - 1, -1, -1)),
        "mode": "markers+text",
        "marker": {"color": "#1f77b4", "size": 10, "symbol": "circle"},
        "text": [_significance_stars(c.get("pvalue", 1.0)) for c in sorted_coefs],
        "textposition": "middle right",
        "textfont": {"size": 12, "color": "red"},
        "name": "Coefficient",
        "showlegend": False,
        "hovertemplate": "%{text}<br>Coeff: %{x:.4f}<extra></extra>",
        "customdata": names,
    })

    layout = {
        "title": {"text": "Coefficient Estimates (Dot-Whisker)", "x": 0.5},
        "xaxis": {
            "title": "Coefficient Estimate",
            "zeroline": True,
            "zerolinecolor": "gray",
            "zerolinewidth": 1,
        },
        "yaxis": {
            "tickvals": list(range(n)),
            "ticktext": list(reversed(names)),
            "title": "",
        },
        "template": "plotly_white",
        "height": max(300, n * 40),
        "annotations": [
            {
                "xref": "paper", "yref": "paper",
                "x": 1, "y": -0.08,
                "text": "*** p<0.01, ** p<0.05, * p<0.1",
                "showarrow": False,
                "font": {"size": 10, "color": "gray"},
                "xanchor": "right",
            }
        ],
    }

    chart_spec = {"data": traces, "layout": layout}
    return json.dumps({"success": True, "chart": chart_spec})


def _make_residual_fitted_chart(residuals: np.ndarray, fitted: np.ndarray) -> dict:
    traces = [
        {
            "type": "scatter",
            "x": fitted.tolist(),
            "y": residuals.tolist(),
            "mode": "markers",
            "marker": {"color": "steelblue", "size": 6, "opacity": 0.6},
            "name": "Residuals",
            "showlegend": False,
            "hovertemplate": "Fitted: %{x:.4f}<br>Residual: %{y:.4f}<extra></extra>",
        },
        {
            "type": "scatter",
            "x": [float(fitted.min()), float(fitted.max())],
            "y": [0, 0],
            "mode": "lines",
            "line": {"color": "red", "width": 1.5, "dash": "dash"},
            "name": "y=0",
            "showlegend": False,
            "hoverinfo": "none",
        },
    ]
    layout = {
        "title": {"text": "Residuals vs Fitted", "x": 0.5},
        "xaxis": {"title": "Fitted Values"},
        "yaxis": {"title": "Residuals"},
        "template": "plotly_white",
    }
    return {"data": traces, "layout": layout}


def _make_qq_chart(residuals: np.ndarray) -> dict:
    n = len(residuals)
    theoretical = np.sort(_norm_ppf(np.arange(1, n + 1) / (n + 1)))
    sample = np.sort(residuals)

    min_val = float(min(theoretical.min(), sample.min()))
    max_val = float(max(theoretical.max(), sample.max()))

    traces = [
        {
            "type": "scatter",
            "x": theoretical.tolist(),
            "y": sample.tolist(),
            "mode": "markers",
            "marker": {"color": "steelblue", "size": 6, "opacity": 0.6},
            "name": "Sample Quantiles",
            "showlegend": False,
            "hovertemplate": "Theoretical: %{x:.4f}<br>Sample: %{y:.4f}<extra></extra>",
        },
        {
            "type": "scatter",
            "x": [min_val, max_val],
            "y": [min_val, max_val],
            "mode": "lines",
            "line": {"color": "red", "width": 1.5, "dash": "dash"},
            "name": "Normal",
            "showlegend": False,
            "hoverinfo": "none",
        },
    ]
    layout = {
        "title": {"text": "Normal Q-Q Plot", "x": 0.5},
        "xaxis": {"title": "Theoretical Quantiles"},
        "yaxis": {"title": "Sample Quantiles"},
        "template": "plotly_white",
    }
    return {"data": traces, "layout": layout}


def _norm_ppf(q: np.ndarray) -> np.ndarray:
    """Normal distribution PPF approximation."""
    try:
        from scipy.stats import norm
        return norm.ppf(q)
    except ImportError:
        pass
    p = np.clip(np.asarray(q, dtype=float), 1e-15, 1 - 1e-15)
    t = np.sqrt(-2 * np.log(1 - p))
    c0, c1, c2 = 2.515517, 0.802853, 0.010328
    d1, d2, d3 = 1.432788, 0.189269, 0.001308
    return t - (c0 + c1 * t + c2 * t**2) / (1 + d1 * t + d2 * t**2 + d3 * t**3)


def _make_scale_location_chart(residuals: np.ndarray, fitted: np.ndarray) -> dict:
    std_resid = residuals / np.std(residuals, ddof=1)
    sqrt_abs = np.sqrt(np.abs(std_resid))

    traces = [
        {
            "type": "scatter",
            "x": fitted.tolist(),
            "y": sqrt_abs.tolist(),
            "mode": "markers",
            "marker": {"color": "steelblue", "size": 6, "opacity": 0.6},
            "name": "std residual",
            "showlegend": False,
            "hovertemplate": "Fitted: %{x:.4f}<br>sqrt(|std res|): %{y:.4f}<extra></extra>",
        },
    ]
    layout = {
        "title": {"text": "Scale-Location Plot", "x": 0.5},
        "xaxis": {"title": "Fitted Values"},
        "yaxis": {"title": "sqrt(|Standardized Residuals|)"},
        "template": "plotly_white",
    }
    return {"data": traces, "layout": layout}


def _make_cooks_chart(residuals: np.ndarray, fitted: np.ndarray,
                      n_params: int) -> dict:
    n = len(residuals)
    p = max(n_params, 1)
    mse = np.var(residuals, ddof=p)
    if mse <= 0:
        cooks_d = np.zeros(n)
    else:
        cooks_d = (residuals ** 2) / (p * mse) * (1.0 / n)
    threshold = 4.0 / n

    obs_idx = list(range(1, n + 1))
    colors = ["red" if d > threshold else "steelblue" for d in cooks_d]

    traces = [
        {
            "type": "bar",
            "x": obs_idx,
            "y": cooks_d.tolist(),
            "marker": {"color": colors, "opacity": 0.7},
            "name": "Cook's D",
            "showlegend": False,
            "hovertemplate": "Obs: %{x}<br>Cook's D: %{y:.4f}<extra></extra>",
        },
        {
            "type": "scatter",
            "x": [1, n],
            "y": [threshold, threshold],
            "mode": "lines",
            "line": {"color": "red", "width": 1.5, "dash": "dash"},
            "name": f"Threshold (4/n) = {threshold:.4f}",
            "showlegend": False,
            "hoverinfo": "skip",
        },
    ]
    layout = {
        "title": {"text": "Cook's Distance", "x": 0.5},
        "xaxis": {"title": "Observation Index"},
        "yaxis": {"title": "Cook's Distance"},
        "template": "plotly_white",
    }
    return {"data": traces, "layout": layout}


def _make_unavailable_chart(title: str, message: str) -> dict:
    """Create a placeholder chart showing a message when a diagnostic is unavailable."""
    layout = {
        "title": {"text": title, "x": 0.5},
        "xaxis": {"visible": False},
        "yaxis": {"visible": False},
        "template": "plotly_white",
        "annotations": [
            {
                "xref": "paper",
                "yref": "paper",
                "x": 0.5,
                "y": 0.5,
                "text": message,
                "showarrow": False,
                "font": {"size": 13, "color": "#666"},
                "xanchor": "center",
            }
        ],
    }
    return {"data": [], "layout": layout}


# ===========================================================================
# 5. Multi-model comparison chart
# ===========================================================================


def compare_models(model_results_json: str) -> str:
    """Generate a coefficient comparison Plotly chart from multiple model results.

    Args:
        model_results_json: JSON string with a list of {name, result} objects,
                            where each 'result' is a regression result dict.

    Returns:
        JSON with success and chart (plotly spec).
    """
    try:
        models = json.loads(model_results_json)
    except json.JSONDecodeError as e:
        return json.dumps({"success": False, "error": f"JSON parse error: {e}"})

    if not models or len(models) < 2:
        return json.dumps({"success": False, "error": "Need at least 2 models to compare."})

    # Collect all unique coefficient names (excluding Intercept) across models
    all_coefs: list[str] = []
    for m in models:
        coefs = m.get("result", {}).get("coefficients", [])
        for c in coefs:
            name = c.get("name", "")
            if name and name != "Intercept" and name not in all_coefs:
                all_coefs.append(name)

    if not all_coefs:
        return json.dumps({"success": False, "error": "No non-intercept coefficients to compare."})

    # Sort by the absolute average coefficient value across models
    # Reversed so largest absolute coef appears at top of plot
    avg_abs = {}
    for name in all_coefs:
        vals = []
        for m in models:
            coefs = m.get("result", {}).get("coefficients", [])
            for c in coefs:
                if c.get("name") == name:
                    vals.append(abs(c.get("coef", 0)))
                    break
        avg_abs[name] = sum(vals) / len(vals) if vals else 0
    all_coefs.sort(key=lambda x: avg_abs.get(x, 0), reverse=True)

    n_coefs = len(all_coefs)
    n_models = len(models)

    colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
              "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf"]

    traces = []
    for mi, model_entry in enumerate(models):
        model_name = model_entry.get("name", f"Model {mi + 1}")
        color = colors[mi % len(colors)]
        coef_dict = {}
        for c in model_entry.get("result", {}).get("coefficients", []):
            coef_dict[c.get("name", "")] = c

        y_positions = []
        estimates = []
        ci_lows = []
        ci_highs = []
        for ci, name in enumerate(all_coefs):
            c = coef_dict.get(name)
            if c:
                y_pos = n_coefs - 1 - ci + (mi - (n_models - 1) / 2) * 0.3
                y_positions.append(y_pos)
                estimates.append(c.get("coef", 0))
                ci_low = c.get("ci_lower")
                ci_high = c.get("ci_upper")
                # Only add CI if both bounds are valid (not None, not both 0)
                has_ci = (ci_low is not None and ci_high is not None and
                          not (ci_low == 0 and ci_high == 0))
                ci_lows.append(ci_low if has_ci else None)
                ci_highs.append(ci_high if has_ci else None)
        if not estimates:
            continue

        # CI whiskers — skip coefficients lacking valid CI data
        for i in range(len(estimates)):
            if ci_lows[i] is not None:
                traces.append({
                    "type": "scatter",
                    "x": [ci_lows[i], ci_highs[i]],
                    "y": [y_positions[i], y_positions[i]],
                    "mode": "lines",
                    "line": {"color": color, "width": 2},
                    "showlegend": False,
                    "hoverinfo": "none",
                })

        # Dot markers
        traces.append({
            "type": "scatter",
            "x": estimates,
            "y": y_positions,
            "mode": "markers",
            "marker": {"color": color, "size": 10, "symbol": "circle"},
            "name": model_name,
            "showlegend": True,
            "hovertemplate": "%{x:.4f}<extra>" + model_name + "</extra>",
        })

    # Y-axis tick labels: coefficient names (one per row)
    tick_vals = list(range(n_coefs))
    tick_texts = list(reversed(all_coefs))

    layout = {
        "title": {"text": "Model Comparison: Coefficient Estimates", "x": 0.5},
        "xaxis": {
            "title": "Coefficient Estimate",
            "zeroline": True,
            "zerolinecolor": "gray",
            "zerolinewidth": 1,
        },
        "yaxis": {
            "tickvals": tick_vals,
            "ticktext": tick_texts,
            "title": "",
        },
        "template": "plotly_white",
        "height": max(300, n_coefs * 45 + 60),
        "legend": {"orientation": "h", "yanchor": "bottom", "y": 1.02, "xanchor": "right", "x": 1},
    }

    chart_spec = {"data": traces, "layout": layout}
    return json.dumps({"success": True, "chart": chart_spec})


# ===========================================================================
# 6. Scatter chart with regression line
# ===========================================================================


def generate_scatter_chart(data_json: str, x_var: str, y_var: str) -> str:
    """Generate a scatter plot with fitted regression line and 95% CI band.

    Args:
        data_json: JSON string with 'data' (list of lists) and 'columns' metadata.
        x_var: Name of the X-axis variable (independent variable).
        y_var: Name of the Y-axis variable (dependent variable).

    Returns:
        JSON with success and chart (plotly spec).
    """
    try:
        data_dict = json.loads(data_json)
    except json.JSONDecodeError as e:
        return json.dumps({"success": False, "error": f"JSON parse error: {e}"})

    # Reconstruct DataFrame
    try:
        if "data" in data_dict and isinstance(data_dict["data"], list):
            rows = data_dict["data"]
            if len(rows) < 2:
                return json.dumps({"success": False, "error": "Data has no rows."})
            headers = rows[0]
            df = pd.DataFrame(rows[1:], columns=headers)
            if _validate_columns_metadata(data_dict.get("columns"), df):
                for col_info in data_dict["columns"]:
                    if (col_info.get("col_type") == "numeric"
                            and isinstance(col_info.get("name"), str)
                            and col_info["name"] in df.columns):
                        df[col_info["name"]] = pd.to_numeric(df[col_info["name"]], errors="coerce")
            else:
                print(
                    "[bridge] columns metadata missing or invalid, "
                    "using dtype inference fallback",
                    file=sys.stderr,
                )
                df = _infer_numeric_columns(df)
        else:
            return json.dumps({"success": False, "error": "Invalid data format."})
    except Exception as e:
        return json.dumps({"success": False, "error": f"DataFrame construction error: {e}"})

    if x_var not in df.columns or y_var not in df.columns:
        return json.dumps({"success": False, "error": "Variable not found in data."})

    # Drop rows with missing values in relevant columns
    df_scatter = df[[x_var, y_var]].dropna()
    if len(df_scatter) < 3:
        return json.dumps({"success": False, "error": "Not enough valid data points (<3)."})

    x = pd.to_numeric(df_scatter[x_var], errors="coerce").values
    y = pd.to_numeric(df_scatter[y_var], errors="coerce").values
    n = len(x)

    # OLS fit for slope/intercept
    x_mean = float(np.mean(x))
    y_mean = float(np.mean(y))
    numerator = float(np.sum((x - x_mean) * (y - y_mean)))
    denominator = float(np.sum((x - x_mean) ** 2))
    if denominator == 0:
        return json.dumps({"success": False, "error": "X variable has zero variance."})

    beta = numerator / denominator
    alpha = y_mean - beta * x_mean

    # Predicted values and residuals
    y_pred = alpha + beta * x
    residuals = y - y_pred
    mse = float(np.sum(residuals ** 2)) / (n - 2) if n > 2 else 0.0

    # Sort for smooth line
    sort_idx = np.argsort(x)
    x_sorted = x[sort_idx]
    y_pred_sorted = y_pred[sort_idx]

    # 95% CI band: se_pred = sqrt(MSE * (1/n + (x_i - x_mean)^2 / SXX))
    SXX = float(np.sum((x - x_mean) ** 2))  # noqa: N806
    if SXX > 0 and mse > 0:
        try:
            from scipy import stats as scipy_stats
            t_crit = float(scipy_stats.t.ppf(0.975, df=n - 2))
        except (ImportError, Exception):
            # Fallback: use normal approximation (z = 1.96)
            t_crit = 1.96
        se_pred_line = np.sqrt(mse * (1.0 / n + (x_sorted - x_mean) ** 2 / SXX))
        ci_low = y_pred_sorted - t_crit * se_pred_line
        ci_high = y_pred_sorted + t_crit * se_pred_line
    else:
        ci_low = y_pred_sorted
        ci_high = y_pred_sorted

    traces = [
        # Scatter points
        {
            "type": "scatter",
            "x": x.tolist(),
            "y": y.tolist(),
            "mode": "markers",
            "marker": {"color": "steelblue", "size": 7, "opacity": 0.5},
            "name": "Observations",
            "showlegend": True,
            "hovertemplate": f"{x_var}: %{{x:.4f}}<br>{y_var}: %{{y:.4f}}<extra></extra>",
        },
        # Regression line
        {
            "type": "scatter",
            "x": x_sorted.tolist(),
            "y": y_pred_sorted.tolist(),
            "mode": "lines",
            "line": {"color": "red", "width": 2},
            "name": f"y = {alpha:.4f} + {beta:.4f}*x",
            "showlegend": True,
            "hovertemplate": "%{y:.4f}<extra></extra>",
        },
        # 95% CI band (upper bound)
        {
            "type": "scatter",
            "x": x_sorted.tolist(),
            "y": ci_high.tolist(),
            "mode": "lines",
            "line": {"color": "gray", "width": 0, "dash": "dash"},
            "name": "95% CI upper",
            "showlegend": False,
            "hoverinfo": "skip",
        },
        # 95% CI band (lower bound with fill)
        {
            "type": "scatter",
            "x": x_sorted.tolist(),
            "y": ci_low.tolist(),
            "mode": "lines",
            "line": {"color": "gray", "width": 0},
            "fill": "tonexty",
            "fillcolor": "rgba(128, 128, 128, 0.2)",
            "name": "95% CI",
            "showlegend": True,
            "hoverinfo": "skip",
        },
    ]

    layout = {
        "title": {"text": f"{y_var} vs {x_var}", "x": 0.5},
        "xaxis": {"title": x_var},
        "yaxis": {"title": y_var},
        "template": "plotly_white",
        "height": 400,
    }

    chart_spec = {"data": traces, "layout": layout}
    return json.dumps({"success": True, "chart": chart_spec})


# ===========================================================================
# 6.5. Logit-specific charts: ROC curve and Odds Ratio forest plot
# ===========================================================================


def generate_roc_chart(result_json: str) -> str:
    """Generate an ROC curve from the already-fitted logit model predictions.

    Uses the fitted_values (predicted probabilities) and y_actual from the
    regression result dict, so the ROC reflects the user's actual model.

    Args:
        result_json: JSON string of a logit model result (from _extract_logit_result).

    Returns:
        JSON with success and chart (plotly spec).
    """
    try:
        result = json.loads(result_json)
    except json.JSONDecodeError as e:
        return json.dumps({"success": False, "error": f"JSON parse error: {e}"})

    if result.get("model_type") not in ("logit", "probit"):
        return json.dumps({
            "success": False,
            "error": "ROC is only available for binary choice models (logit/probit).",
        })

    y_pred_prob = np.array(result.get("fitted_values", []))
    y_actual = np.array(result.get("y_actual", []))

    if len(y_pred_prob) < 5 or len(y_actual) < 5:
        return json.dumps({
            "success": False,
            "error": "Not enough valid observations for ROC (<5).",
        })

    if len(y_pred_prob) != len(y_actual):
        return json.dumps({
            "success": False,
            "error": "Mismatch between predictions and actual values.",
        })

    y_unique = np.unique(y_actual)
    if len(y_unique) != 2:
        return json.dumps({
            "success": False,
            "error": f"ROC requires binary response. Found {len(y_unique)} unique values."
        })

    # Code y as 0/1 (handle non-standard binary encodings like -1/+1 or string values)
    y_binary = (y_actual == y_unique[1]).astype(float)

    # Compute ROC curve from the already-fitted predicted probabilities
    thresholds = np.sort(np.unique(y_pred_prob))[::-1]
    tpr_list = []
    fpr_list = []

    for thr in thresholds:
        y_pred_class = (y_pred_prob >= thr).astype(int)
        tp = int(np.sum((y_pred_class == 1) & (y_binary == 1)))
        fp = int(np.sum((y_pred_class == 1) & (y_binary == 0)))
        fn = int(np.sum((y_pred_class == 0) & (y_binary == 1)))
        tn = int(np.sum((y_pred_class == 0) & (y_binary == 0)))

        tpr = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        fpr = fp / (fp + tn) if (fp + tn) > 0 else 0.0
        tpr_list.append(tpr)
        fpr_list.append(fpr)

    # Compute AUC with trapezoidal rule
    # Sort by FPR ascending (standard ROC)
    sorted_pairs = sorted(zip(fpr_list, tpr_list), key=lambda p: p[0])
    auc = 0.0
    for i in range(1, len(sorted_pairs)):
        fpr_prev, tpr_prev = sorted_pairs[i - 1]
        fpr_curr, tpr_curr = sorted_pairs[i]
        auc += (fpr_curr - fpr_prev) * (tpr_prev + tpr_curr) / 2.0

    # Build Plotly chart
    fpr_sorted = [p[0] for p in sorted_pairs]
    tpr_sorted = [p[1] for p in sorted_pairs]

    traces = [
        {
            "type": "scatter",
            "x": fpr_sorted,
            "y": tpr_sorted,
            "mode": "lines",
            "line": {"color": "steelblue", "width": 2},
            "name": f"ROC (AUC = {auc:.4f})",
            "hovertemplate": "FPR: %{x:.4f}<br>TPR: %{y:.4f}<extra></extra>",
        },
        {
            "type": "scatter",
            "x": [0, 1],
            "y": [0, 1],
            "mode": "lines",
            "line": {"color": "gray", "width": 1.5, "dash": "dash"},
            "name": "Random Classifier",
            "showlegend": True,
            "hoverinfo": "none",
        },
    ]

    layout = {
        "title": {"text": f"ROC Curve (AUC = {auc:.4f})", "x": 0.5},
        "xaxis": {"title": "False Positive Rate (1 - Specificity)", "range": [0, 1]},
        "yaxis": {"title": "True Positive Rate (Sensitivity)", "range": [0, 1]},
        "template": "plotly_white",
        "height": 400,
        "showlegend": True,
        "legend": {"x": 0.6, "y": 0.1},
    }

    chart_spec = {"data": traces, "layout": layout}
    return json.dumps({"success": True, "chart": chart_spec, "auc": round(auc, 4)})


def generate_or_chart(result_json: str) -> str:
    """Generate an odds ratio forest plot from logit regression results.

    Args:
        result_json: JSON string of a logit model result.

    Returns:
        JSON with success and chart (plotly spec).
    """
    try:
        result = json.loads(result_json)
    except json.JSONDecodeError as e:
        return json.dumps({"success": False, "error": f"JSON parse error: {e}"})

    coefficients = result.get("coefficients", [])
    if not coefficients:
        return json.dumps({"success": False, "error": "No coefficients."})

    coeffs = [c for c in coefficients if c.get("coef") != 0 or c["name"] == "Intercept"]
    if not coeffs:
        return json.dumps({"success": False, "error": "No valid coefficients."})

    # Separate intercept and sort others by OR magnitude
    intercept = None
    others = []
    for c in coeffs:
        if c["name"] == "Intercept":
            intercept = c
        else:
            others.append(c)
    others.sort(key=lambda x: abs(x.get("odds_ratio", 1) - 1), reverse=True)

    sorted_coefs = others  # Skip intercept for OR plot
    if intercept:
        sorted_coefs = [intercept] + sorted_coefs

    n = len(sorted_coefs)
    names = [c["name"] for c in sorted_coefs]
    or_vals = [c.get("odds_ratio", float(np.exp(c["coef"]))) for c in sorted_coefs]
    or_lows = [c.get("or_ci_lower", float(np.exp(c["ci_lower"]))) for c in sorted_coefs]
    or_highs = [c.get("or_ci_upper", float(np.exp(c["ci_upper"]))) for c in sorted_coefs]

    # Limit OR range for display
    max_or = max(max(or_highs), 3.0)

    traces = []
    for i in range(n):
        traces.append({
            "type": "scatter",
            "x": [or_lows[i], or_highs[i]],
            "y": [n - 1 - i, n - 1 - i],
            "mode": "lines",
            "line": {"color": "#1f77b4", "width": 2},
            "showlegend": False,
            "hoverinfo": "none",
        })

    traces.append({
        "type": "scatter",
        "x": or_vals,
        "y": list(range(n - 1, -1, -1)),
        "mode": "markers",
        "marker": {"color": "#1f77b4", "size": 10, "symbol": "circle"},
        "name": "Odds Ratio",
        "showlegend": False,
        "hovertemplate": "OR: %{x:.4f}<br>%{customdata}<extra></extra>",
        "customdata": [
            f"OR={v:.4f} 95%CI [{lo:.4f}, {hi:.4f}]"
            f" {_significance_stars(c.get('pvalue', 1))}"
            for v, lo, hi, c in zip(or_vals, or_lows, or_highs, sorted_coefs)
        ],
    })

    # Reference line at OR = 1
    traces.append({
        "type": "scatter",
        "x": [1, 1],
        "y": [-0.5, n - 0.5],
        "mode": "lines",
        "line": {"color": "red", "width": 1.5, "dash": "dash"},
        "name": "OR = 1",
        "showlegend": True,
        "hoverinfo": "none",
    })

    # Build annotations with OR values on the right side
    annotations = []
    for i, (name, or_v, or_l, or_h) in enumerate(zip(names, or_vals, or_lows, or_highs)):
        stars = _significance_stars(sorted_coefs[i].get("pvalue", 1))
        annotations.append({
            "xref": "paper",
            "yref": "y",
            "x": 1.02,
            "y": n - 1 - i,
            "text": f"OR={or_v:.4f} [{or_l:.4f}, {or_h:.4f}] {stars}",
            "showarrow": False,
            "font": {"size": 9, "color": "#333"},
            "xanchor": "left",
        })

    layout = {
        "title": {"text": "Odds Ratio Forest Plot", "x": 0.5},
        "xaxis": {
            "title": "Odds Ratio (log scale)",
            "type": "log",
            "range": [np.log10(max(0.01, min(min(or_lows) * 0.5, 0.5))), np.log10(max_or * 1.5)],
            "tickformat": ".4f",
        },
        "yaxis": {
            "tickvals": list(range(n)),
            "ticktext": list(reversed(names)),
            "title": "",
        },
        "template": "plotly_white",
        "height": max(300, n * 45 + 60),
        "margin": {"r": 250},
        "annotations": annotations
                     + [{
                         "xref": "paper", "yref": "paper",
                         "x": 1, "y": -0.08,
                         "text": "*** p<0.01, ** p<0.05, * p<0.1",
                         "showarrow": False,
                         "font": {"size": 10, "color": "gray"},
                         "xanchor": "right",
                     }],
    }

    chart_spec = {"data": traces, "layout": layout}
    return json.dumps({"success": True, "chart": chart_spec})


# ===========================================================================
# 7. Export
# ===========================================================================


def export_csv(result_json: str) -> str:
    """Generate a CSV string of the coefficient table."""
    try:
        result = json.loads(result_json)
    except json.JSONDecodeError:
        return json.dumps({"success": False, "error": "Invalid JSON."})

    coefficients = result.get("coefficients", [])
    if not coefficients:
        return json.dumps({"success": False, "error": "No coefficients to export."})

    model_type = result.get("model_type", "")
    is_logit = model_type == "logit"
    is_probit = model_type == "probit"
    is_count = model_type in ("poisson", "negbin")
    is_mle = is_logit or is_probit or is_count
    stat_col = "z-value" if is_mle else "t-value"
    extra_col = ""
    extra_field = ""
    if is_logit:
        extra_col = ",Odds Ratio"
        extra_field = "odds_ratio"
    elif is_count:
        extra_col = ",IRR"
        extra_field = "irr"

        extra_header = f"{extra_col},Significance"
    header = (
        f"Variable,Coefficient,Std.Err.,{stat_col},p-value,"
        f"CI(95%) Low,CI(95%) High{extra_header}"
    )
    lines = [header]
    for c in coefficients:
        stat_val = c.get("z_stat", c.get("t_stat", 0))
        extra_val = f',{c.get(extra_field, "")}' if extra_field else ""
        lines.append(
            f'"{c["name"]}",{c["coef"]},{c["se"]},{stat_val},'
            f'{c["pvalue"]},{c["ci_lower"]},{c["ci_upper"]}{extra_val},{c["significance"]}'
        )

    csv_text = "\n".join(lines)
    if is_mle:
        subtype = model_type.upper()
        model_info = (
            f"\n\n# Model Summary ({subtype})\n"
            f'# Model Type,{subtype}\n'
            f'# Pseudo R-squared,{result.get("pseudo_r_squared", "N/A")}\n'
            f'# LR chi2,{result.get("llr", "N/A")}\n'
            f'# LR p-value,{result.get("llr_pvalue", "N/A")}\n'
            f'# Log-Likelihood,{result.get("log_likelihood", "N/A")}\n'
            f'# AIC,{result.get("aic", "N/A")}\n'
            f'# BIC,{result.get("bic", "N/A")}\n'
            f'# N,{result.get("n_obs", "N/A")}\n'
            f'# Specification,"{result.get("specification", "")}"\n'
        )
    else:
        model_info = (
            f"\n\n# Model Summary\n"
            f'# Model Type,{model_type.upper()}\n'
            f'# R-squared,{result.get("r_squared", "N/A")}\n'
            f'# Adj R-squared,{result.get("adj_r_squared", "N/A")}\n'
            f'# RMSE,{result.get("rmse", "N/A")}\n'
            f'# AIC,{result.get("aic", "N/A")}\n'
            f'# BIC,{result.get("bic", "N/A")}\n'
            f'# N,{result.get("n_obs", "N/A")}\n'
            f'# Specification,"{result.get("specification", "")}"\n'
        )

    return json.dumps({"success": True, "csv": csv_text + model_info})


def export_excel(result_json: str) -> str:
    """Export as Excel (base64-encoded) or fallback to CSV."""
    try:
        result = json.loads(result_json)
    except json.JSONDecodeError:
        return json.dumps({"success": False, "error": "Invalid JSON."})

    import base64

    try:
        import openpyxl  # noqa: F401
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Regression Results"

        is_logit = result.get("model_type", "") == "logit"
        is_probit = result.get("model_type", "") == "probit"
        is_count = result.get("model_type", "") in ("poisson", "negbin")
        is_mle = is_logit or is_probit or is_count  # noqa: F841

        # Title
        title_text = "Logit Regression Results" if is_logit else "OLS Regression Results"
        ws.merge_cells("A1:I1") if is_logit else ws.merge_cells("A1:H1")
        ws["A1"] = title_text
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Model info
        if is_logit:
            info_items = [
                ("Dependent Variable", result.get("dep_var", "")),
                ("Specification", result.get("specification", "")),
                ("Method", "Logit (MLE)"),
                ("N", result.get("n_obs", "")),
                ("Pseudo R-squared", result.get("pseudo_r_squared", "")),
                ("LR chi2", result.get("llr", "")),
                ("LR p-value", result.get("llr_pvalue", "")),
                ("Log-Likelihood", result.get("log_likelihood", "")),
                ("AIC", result.get("aic", "")),
                ("BIC", result.get("bic", "")),
            ]
        else:
            info_items = [
                ("Dependent Variable", result.get("dep_var", "")),
                ("Specification", result.get("specification", "")),
                ("N", result.get("n_obs", "")),
                ("R-squared", result.get("r_squared", "")),
                ("Adj R-squared", result.get("adj_r_squared", "")),
                ("RMSE", result.get("rmse", "")),
                ("AIC", result.get("aic", "")),
                ("BIC", result.get("bic", "")),
                ("F-statistic",
                 f'{result["f_statistic"][0] if result.get("f_statistic") else "N/A"}'),
            ]
        for i, (label, value) in enumerate(info_items, start=3):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)

        # Coefficient table header
        start_row = len(info_items) + 5
        if is_logit:
            headers = ["Variable", "Coefficient", "Std. Error", "z-value",
                       "Odds Ratio", "p-value", "CI Low (95%)", "CI High (95%)", "Significance"]
        else:
            headers = ["Variable", "Coefficient", "Std. Error", "t-value",
                       "p-value", "CI Low (95%)", "CI High (95%)", "Significance"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=j, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Coefficient rows
        n_columns = len(headers)
        for i, c in enumerate(result.get("coefficients", [])):
            row = start_row + 1 + i
            stat_val = c.get("z_stat", c.get("t_stat", 0))
            if is_logit:
                ws.cell(row=row, column=1, value=c["name"])
                ws.cell(row=row, column=2, value=round(c["coef"], 6))
                ws.cell(row=row, column=3, value=round(c["se"], 6))
                ws.cell(row=row, column=4, value=round(stat_val, 4))
                ws.cell(row=row, column=5, value=round(c.get("odds_ratio", 0), 6))
                ws.cell(row=row, column=6, value=c["pvalue"])
                ws.cell(row=row, column=7, value=round(c["ci_lower"], 6))
                ws.cell(row=row, column=8, value=round(c["ci_upper"], 6))
                ws.cell(row=row, column=9, value=c["significance"])
            else:
                ws.cell(row=row, column=1, value=c["name"])
                ws.cell(row=row, column=2, value=round(c["coef"], 6))
                ws.cell(row=row, column=3, value=round(c["se"], 6))
                ws.cell(row=row, column=4, value=round(stat_val, 4))
                ws.cell(row=row, column=5, value=c["pvalue"])
                ws.cell(row=row, column=6, value=round(c["ci_lower"], 6))
                ws.cell(row=row, column=7, value=round(c["ci_upper"], 6))
                ws.cell(row=row, column=8, value=c["significance"])

            # Highlight significant rows
            if c.get("pvalue", 1) < 0.05:
                for j in range(1, n_columns + 1):
                    ws.cell(row=row, column=j).fill = PatternFill(
                        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
                    )

        # Column widths
        ws.column_dimensions["A"].width = 20
        if is_logit:
            for col in "BCDEFGHI":
                ws.column_dimensions[col].width = 16
        else:
            for col in "BCDEFGH":
                ws.column_dimensions[col].width = 16

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return json.dumps({
            "success": True,
            "excel_b64": base64.b64encode(buf.read()).decode("ascii"),
            "filename": "regression_results.xlsx",
        })
    except ImportError:
        # Fall back to CSV export
        return export_csv(result_json)


# ===========================================================================
# 8. Gallery data access (pre-computed results as JSON)
# ===========================================================================


def get_gallery_index() -> str:
    """Return gallery index (lightweight metadata, no data)."""
    items = [
        {
            "id": "survey_happiness",
            "title": "CGSS Resident Happiness Survey",
            "persona": "Social Science Grad Student",
            "description": (
                "Study income, education, health, urban/rural status, "
                "and work hours on subjective well-being."
            ),
            "tags": ["Survey Data", "Categorical", "Multicollinearity"],
            "n_obs": 400,
            "dep_var": "happiness_score",
        },
        {
            "id": "trust_experiment",
            "title": "Social Trust Determinants",
            "persona": "Social Science Grad Student",
            "description": (
                "200-sample survey on age, income, education, "
                "media exposure, and party membership effects on trust."
            ),
            "tags": ["Small Sample", "Borderline Significance", "Social Survey"],
            "n_obs": 200,
            "dep_var": "trust_index",
        },
        {
            "id": "ecommerce_sales",
            "title": "E-commerce Sales Drivers",
            "persona": "Market Researcher",
            "description": (
                "500 days of e-commerce data: ad spend, price, promotions, "
                "competitor price, and season effects on sales."
            ),
            "tags": ["Business Analytics", "High R-squared", "Multicollinearity"],
            "n_obs": 500,
            "dep_var": "sales",
        },
        {
            "id": "customer_satisfaction",
            "title": "Restaurant Customer Satisfaction",
            "persona": "Market Researcher",
            "description": (
                "350 surveys analyzing wait time, service quality, "
                "price perception, loyalty, and complaints."
            ),
            "tags": ["Customer Analysis", "Multi-category", "Service Industry"],
            "n_obs": 350,
            "dep_var": "satisfaction_score",
        },
        {
            "id": "policy_effect",
            "title": "Environmental Policy Evaluation",
            "persona": "Policy Analyst",
            "description": (
                "300 city-level data on environmental regulation intensity, "
                "GDP, industrial structure, and emission reduction."
            ),
            "tags": ["Policy Evaluation", "Interaction Terms", "Robust SE"],
            "n_obs": 300,
            "dep_var": "emission_reduction",
        },
    ]
    return json.dumps(items)
